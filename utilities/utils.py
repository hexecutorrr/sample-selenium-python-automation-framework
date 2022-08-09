import inspect
import logging
from pathlib import Path
from openpyxl import load_workbook
import csv


class Utils:

    @staticmethod
    def path():
        return str(Path(__file__).parent.parent.absolute())

    @staticmethod
    def custom_logger(log_level=logging.DEBUG):
        log_path = Utils.path()+"\\logs\\automation.log"
        # Clear log file before writing
        clear_log = open(log_path, 'r+')
        clear_log.truncate(0)
        clear_log.close()
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # Create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(log_level)
        # Create console handler
        fh = logging.FileHandler(filename=log_path)
        # Create log formatter
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%m/%d/%Y %H:%M:%S')
        # Add formatter to console
        fh.setFormatter(formatter)
        # Add console handler to logger
        logger.addHandler(fh)
        return logger

    @staticmethod
    def read_from_excel(file_name, sheet):
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        data_list = []
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list

    @staticmethod
    def read_from_csv(file_name):
        data = open(file_name, "r")
        csv_reader = csv.reader(data)
        next(csv_reader)
        data_list = []
        for row in csv_reader:
            data_list.append(row)
        return data_list
